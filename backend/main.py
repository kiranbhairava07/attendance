"""
Attendance System — FastAPI Backend
Fixed & production-ready
"""
import dotenv
from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import Optional
from datetime import datetime, timedelta, date, timezone
import asyncpg
import os
from jose import JWTError, jwt
from passlib.context import CryptContext
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import math
import pytz
#import dotenv
dotenv.load_dotenv()  # Load .env file if present

# ─── Config ───────────────────────────────────────────────
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5432/attendance_db"
)
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = int(os.getenv("ACCESS_TOKEN_EXPIRE_HOURS", "8"))

# Office timezone — all shift comparisons done in local time
# Override via env var if needed: e.g. "America/New_York"
OFFICE_TIMEZONE = os.getenv("OFFICE_TIMEZONE", "Asia/Kolkata")

# ─── App ──────────────────────────────────────────────────
app = FastAPI(title="Attendance System")
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Database ─────────────────────────────────────────────
db_pool: Optional[asyncpg.Pool] = None


@app.on_event("startup")
async def startup():
    global db_pool
    db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=5, max_size=20)


@app.on_event("shutdown")
async def shutdown():
    if db_pool:
        await db_pool.close()


async def get_db():
    async with db_pool.acquire() as conn:
        yield conn


# ─── Models ───────────────────────────────────────────────
class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class PunchRequest(BaseModel):
    latitude: float
    longitude: float


# ─── Helpers ──────────────────────────────────────────────
def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def create_token(user_id: int, email: str, role: str) -> str:
    expire = datetime.utcnow() + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    return jwt.encode(
        {"sub": str(user_id), "email": email, "role": role, "exp": expire},
        SECRET_KEY,
        algorithm=ALGORITHM,
    )


def get_local_now() -> datetime:
    """Return current datetime in office timezone (aware)."""
    tz = pytz.timezone(OFFICE_TIMEZONE)
    return datetime.now(tz)


def utc_to_local(dt: datetime) -> datetime:
    """Convert a UTC-aware or naive datetime to office local time."""
    if dt is None:
        return None
    tz = pytz.timezone(OFFICE_TIMEZONE)
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    return dt.astimezone(tz)


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: asyncpg.Connection = Depends(get_db),
):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        user_id = int(user_id)
    except (JWTError, ValueError):
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    user = await db.fetchrow(
        """SELECT u.id, u.email, u.full_name, u.role, u.branch_id,
                  u.shift_start, u.shift_end,
                  b.name as branch_name, b.city as branch_city,
                  b.latitude as branch_lat, b.longitude as branch_lng,
                  b.radius_meters
           FROM users u
           LEFT JOIN branches b ON u.branch_id = b.id
           WHERE u.id = $1 AND u.is_active = TRUE""",
        user_id,
    )

    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    return dict(user)


async def require_hr(user: dict = Depends(get_current_user)):
    if user["role"] not in ["hr", "admin"]:
        raise HTTPException(status_code=403, detail="HR access required")
    return user


# ══════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════

@app.post("/api/auth/register")
async def register(req: LoginRequest, db: asyncpg.Connection = Depends(get_db)):
    existing = await db.fetchrow("SELECT id FROM users WHERE email = $1", req.email.lower())
    if existing:
        raise HTTPException(status_code=409, detail="Email already registered")

    user = await db.fetchrow(
        """INSERT INTO users (email, password_hash, full_name, role, is_active)
           VALUES ($1, $2, $3, 'employee', TRUE)
           RETURNING id, email, full_name, role""",
        req.email.lower(),
        hash_password(req.password),
        req.email.split("@")[0].replace(".", " ").title(),
    )
    return dict(user)


@app.post("/api/auth/login")
async def login(req: LoginRequest, db: asyncpg.Connection = Depends(get_db)):
    user = await db.fetchrow(
        "SELECT id, email, password_hash, full_name, role, is_active FROM users WHERE email = $1",
        req.email.lower(),
    )

    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if not user["is_active"]:
        raise HTTPException(status_code=403, detail="Account deactivated")

    await db.execute("UPDATE users SET last_login = NOW() WHERE id = $1", user["id"])

    token = create_token(user["id"], user["email"], user["role"])
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "email": user["email"],
            "full_name": user["full_name"],
            "role": user["role"],
        },
    }


@app.post("/api/auth/logout")
async def logout():
    return {"message": "Logged out successfully"}


@app.get("/api/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    # Serialize time fields to HH:MM strings for the frontend
    result = dict(user)
    if result.get("shift_start"):
        result["shift_start"] = result["shift_start"].strftime("%H:%M")
    if result.get("shift_end"):
        result["shift_end"] = result["shift_end"].strftime("%H:%M")
    # Serialize decimal coords to float
    for field in ("branch_lat", "branch_lng"):
        if result.get(field) is not None:
            result[field] = float(result[field])
    return result


# ══════════════════════════════════════════════════════════
# ATTENDANCE ROUTES
# ══════════════════════════════════════════════════════════

@app.post("/api/attendance/punch-in")
async def punch_in(
    req: PunchRequest,
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    if not user["branch_id"]:
        raise HTTPException(status_code=400, detail="No branch assigned. Contact HR.")

    # ── Geofence check ──────────────────────────────────────
    distance = haversine(
        req.latitude, req.longitude,
        float(user["branch_lat"]), float(user["branch_lng"])
    )
    radius = user["radius_meters"] or 200
    if distance > radius:
        raise HTTPException(
            status_code=403,
            detail=f"You are {int(distance)}m from office. Must be within {radius}m to punch in."
        )

    # ── One punch-in per day — block if complete cycle exists ──
    today_logs = await db.fetch(
        """SELECT punch_type FROM attendance_logs
           WHERE user_id = $1 AND (punched_at AT TIME ZONE $2)::date = (NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at ASC""",
        user["id"], OFFICE_TIMEZONE
    )

    punch_types = [r["punch_type"] for r in today_logs]

    # Already has an 'in' without 'out' → already punched in
    if punch_types and punch_types[-1] == "in":
        raise HTTPException(status_code=409, detail="Already punched in. Please punch out first.")

    # Completed a full in→out cycle → cannot punch in again today
    if "in" in punch_types and "out" in punch_types:
        raise HTTPException(
            status_code=409,
            detail="You have already completed your attendance for today. See you tomorrow!"
        )

    # ── Calculate late status using LOCAL time ───────────────
    local_now = get_local_now()
    punch_time_local = local_now.time()
    shift_start = user["shift_start"]  # already a time object from DB

    is_late = punch_time_local > shift_start
    late_minutes = 0
    if is_late:
        delta = datetime.combine(date.today(), punch_time_local) - datetime.combine(date.today(), shift_start)
        late_minutes = max(0, int(delta.total_seconds() / 60))

    async with db.transaction():
        log = await db.fetchrow(
            """INSERT INTO attendance_logs
               (user_id, branch_id, punch_type, latitude, longitude, distance_meters, is_valid)
               VALUES ($1, $2, 'in', $3, $4, $5, TRUE)
               RETURNING id, punched_at""",
            user["id"], user["branch_id"],
            req.latitude, req.longitude, int(distance)
        )

        # Insert/update daily summary — status is 'present' (late is tracked separately)
        await db.execute(
            """INSERT INTO daily_summary
               (user_id, work_date, first_punch_in, is_late, late_by_minutes, status)
               VALUES ($1, (NOW() AT TIME ZONE $2)::date, $3, $4, $5, 'present')
               ON CONFLICT (user_id, work_date) DO UPDATE SET
                 first_punch_in   = EXCLUDED.first_punch_in,
                 is_late          = EXCLUDED.is_late,
                 late_by_minutes  = EXCLUDED.late_by_minutes,
                 status           = 'present'""",
            user["id"], OFFICE_TIMEZONE, log["punched_at"], is_late, late_minutes
        )

    local_punch = utc_to_local(log["punched_at"])
    return {
        "success": True,
        "message": "Punched in successfully" + (" — but you're late by " + str(late_minutes) + " min" if is_late else ""),
        "time": local_punch.strftime("%I:%M %p"),
        "distance": int(distance),
        "is_late": is_late,
        "late_by_minutes": late_minutes,
    }


@app.post("/api/attendance/punch-out")
async def punch_out(
    req: PunchRequest,
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    if not user["branch_id"]:
        raise HTTPException(status_code=400, detail="No branch assigned. Contact HR.")

    # ── Punch-out: record GPS but do NOT enforce geofence ────
    # (employee may leave and then realize they forgot to punch out)
    distance = 0
    if user["branch_lat"] and user["branch_lng"]:
        distance = haversine(
            req.latitude, req.longitude,
            float(user["branch_lat"]), float(user["branch_lng"])
        )

    # ── Must have punched in today ───────────────────────────
    today_logs = await db.fetch(
        """SELECT punch_type FROM attendance_logs
           WHERE user_id = $1 AND (punched_at AT TIME ZONE $2)::date = (NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at ASC""",
        user["id"], OFFICE_TIMEZONE
    )
    punch_types = [r["punch_type"] for r in today_logs]

    if not punch_types or punch_types[-1] != "in":
        raise HTTPException(status_code=409, detail="Must punch in before punching out.")

    # Already completed cycle — shouldn't reach here normally
    if punch_types.count("out") >= 1:
        raise HTTPException(
            status_code=409,
            detail="You have already punched out today."
        )

    async with db.transaction():
        log = await db.fetchrow(
            """INSERT INTO attendance_logs
               (user_id, branch_id, punch_type, latitude, longitude, distance_meters, is_valid)
               VALUES ($1, $2, 'out', $3, $4, $5, TRUE)
               RETURNING id, punched_at""",
            user["id"], user["branch_id"],
            req.latitude, req.longitude, int(distance)
        )

        # Calculate total work minutes from first_punch_in
        summary = await db.fetchrow(
            "SELECT first_punch_in FROM daily_summary WHERE user_id = $1 AND work_date = (NOW() AT TIME ZONE $2)::date",
            user["id"], OFFICE_TIMEZONE
        )

        total_minutes = 0
        if summary and summary["first_punch_in"]:
            delta = log["punched_at"] - summary["first_punch_in"]
            total_minutes = max(0, int(delta.total_seconds() / 60))

        await db.execute(
            """UPDATE daily_summary
               SET last_punch_out = $2, total_minutes = $3
               WHERE user_id = $1 AND work_date = (NOW() AT TIME ZONE $4)::date""",
            user["id"], log["punched_at"], total_minutes, OFFICE_TIMEZONE
        )

    local_punch = utc_to_local(log["punched_at"])
    return {
        "success": True,
        "message": "Punched out successfully. Have a great day!",
        "time": local_punch.strftime("%I:%M %p"),
        "total_hours": f"{total_minutes // 60}h {total_minutes % 60}m",
    }


@app.get("/api/attendance/status")
async def get_status(
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db)
):
    today_logs = await db.fetch(
        """SELECT punch_type, punched_at FROM attendance_logs
           WHERE user_id = $1 AND (punched_at AT TIME ZONE $2)::date = (NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at ASC""",
        user["id"], OFFICE_TIMEZONE
    )

    summary = await db.fetchrow(
        "SELECT * FROM daily_summary WHERE user_id = $1 AND work_date = (NOW() AT TIME ZONE $2)::date",
        user["id"], OFFICE_TIMEZONE
    )

    punch_types = [r["punch_type"] for r in today_logs]
    last = today_logs[-1] if today_logs else None

    # Determine state:
    # - "none"      → no punches yet today
    # - "punched_in"  → last punch is 'in'
    # - "completed"   → full in→out cycle done
    state = "none"
    if punch_types:
        if punch_types[-1] == "in":
            state = "punched_in"
        elif "in" in punch_types and punch_types[-1] == "out":
            state = "completed"

    summary_dict = None
    if summary:
        s = dict(summary)
        # Convert timestamps to local time strings for frontend
        if s.get("first_punch_in"):
            s["first_punch_in"] = utc_to_local(s["first_punch_in"]).strftime("%I:%M %p")
        if s.get("last_punch_out"):
            s["last_punch_out"] = utc_to_local(s["last_punch_out"]).strftime("%I:%M %p")
        summary_dict = s

    return {
        "is_punched_in": state == "punched_in",
        "state": state,          # "none" | "punched_in" | "completed"
        "last_punch": {"punch_type": last["punch_type"], "punched_at": str(last["punched_at"])} if last else None,
        "summary": summary_dict,
    }


@app.get("/api/attendance/today")
async def get_today_logs(
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db)
):
    logs = await db.fetch(
        """SELECT punch_type, punched_at, distance_meters
           FROM attendance_logs
           WHERE user_id = $1 AND (punched_at AT TIME ZONE $2)::date = (NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at ASC""",
        user["id"], OFFICE_TIMEZONE
    )
    result = []
    for log in logs:
        d = dict(log)
        d["punched_at_local"] = utc_to_local(d["punched_at"]).strftime("%I:%M %p")
        result.append(d)
    return result


# ══════════════════════════════════════════════════════════
# HR ROUTES
# ══════════════════════════════════════════════════════════

@app.get("/api/hr/daily-report")
async def daily_report(
    date_str: str = None,
    branch_id: int = None,
    user: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    target_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()

    query = """
        SELECT
            u.id, u.email, u.full_name, u.shift_start, u.shift_end,
            b.name as branch_name, b.city,
            s.first_punch_in, s.last_punch_out, s.total_minutes,
            s.is_late, s.late_by_minutes, COALESCE(s.status, 'absent') as status
        FROM users u
        LEFT JOIN branches b ON u.branch_id = b.id
        LEFT JOIN daily_summary s ON s.user_id = u.id AND s.work_date = $1
        WHERE u.is_active = TRUE AND u.role = 'employee'
    """
    params = [target_date]

    if branch_id:
        query += " AND u.branch_id = $2"
        params.append(branch_id)

    query += " ORDER BY b.name, u.full_name"
    rows = await db.fetch(query, *params)

    employees = []
    for row in rows:
        e = dict(row)
        if e.get("first_punch_in"):
            e["first_punch_in"] = utc_to_local(e["first_punch_in"]).isoformat()
        if e.get("last_punch_out"):
            e["last_punch_out"] = utc_to_local(e["last_punch_out"]).isoformat()
        if e.get("shift_start"):
            e["shift_start"] = e["shift_start"].strftime("%H:%M")
        if e.get("shift_end"):
            e["shift_end"] = e["shift_end"].strftime("%H:%M")
        employees.append(e)

    stats = {
        "total": len(employees),
        "present": sum(1 for e in employees if e["status"] == "present"),
        "absent": sum(1 for e in employees if e["status"] == "absent"),
        "late": sum(1 for e in employees if e.get("is_late")),
    }

    return {"date": target_date.isoformat(), "stats": stats, "employees": employees}


@app.get("/api/hr/branches")
async def get_branches(db: asyncpg.Connection = Depends(get_db)):
    branches = await db.fetch(
        "SELECT id, name, city, address, latitude, longitude, radius_meters FROM branches WHERE is_active = TRUE ORDER BY city, name"
    )
    result = []
    for b in branches:
        d = dict(b)
        d["latitude"] = float(d["latitude"])
        d["longitude"] = float(d["longitude"])
        result.append(d)
    return result


@app.get("/api/hr/export")
async def export_excel(
    date_str: str,
    branch_id: int = None,
    user: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    target_date = datetime.strptime(date_str, "%Y-%m-%d").date()

    query = """
        SELECT u.email, u.full_name, b.name as branch_name, b.city,
               s.first_punch_in, s.last_punch_out, s.total_minutes,
               s.is_late, s.late_by_minutes, COALESCE(s.status, 'absent') as status
        FROM users u
        LEFT JOIN branches b ON u.branch_id = b.id
        LEFT JOIN daily_summary s ON s.user_id = u.id AND s.work_date = $1
        WHERE u.is_active = TRUE AND u.role = 'employee'
    """
    params = [target_date]
    if branch_id:
        query += " AND u.branch_id = $2"
        params.append(branch_id)
    query += " ORDER BY b.name, u.full_name"

    rows = await db.fetch(query, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Email", "Name", "Branch", "City", "Punch In", "Punch Out", "Hours", "Status", "Late By"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B63F6", end_color="3B63F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        status = row["status"]
        fill_color = "ECFDF5" if status == "present" else "FEF2F2"

        pin = utc_to_local(row["first_punch_in"]).strftime("%I:%M %p") if row["first_punch_in"] else "—"
        pout = utc_to_local(row["last_punch_out"]).strftime("%I:%M %p") if row["last_punch_out"] else "—"
        hours = f"{row['total_minutes'] // 60}h {row['total_minutes'] % 60}m" if row["total_minutes"] else "—"

        ws.cell(row_idx, 1, row["email"])
        ws.cell(row_idx, 2, row["full_name"])
        ws.cell(row_idx, 3, row["branch_name"] or "—")
        ws.cell(row_idx, 4, row["city"] or "—")
        ws.cell(row_idx, 5, pin)
        ws.cell(row_idx, 6, pout)
        ws.cell(row_idx, 7, hours)
        ws.cell(row_idx, 8, ("LATE — " if row["is_late"] else "") + status.upper())
        ws.cell(row_idx, 9, f"+{row['late_by_minutes']}m" if row["is_late"] else "On Time")

        for col in range(1, 10):
            ws.cell(row_idx, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col, width in zip("ABCDEFGHI", [26, 22, 26, 16, 12, 12, 10, 14, 10]):
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{target_date.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════
# HEALTH
# ══════════════════════════════════════════════════════════

@app.get("/")
async def root():
    return {"status": "ok", "app": "Attendance System"}


@app.get("/health")
async def health():
    tz = pytz.timezone(OFFICE_TIMEZONE)
    return {
        "status": "healthy",
        "utc": datetime.utcnow().isoformat(),
        "local": datetime.now(tz).isoformat(),
        "timezone": OFFICE_TIMEZONE,
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)